import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as ExcelImage

# =========================================================
# CONFIGURATION
# =========================================================

INPUT_FILE = "sales_data.xlsx"
OUTPUT_FILE = "automated_profitability_analysis.xlsx"

# =========================================================
# LOAD DATA
# =========================================================

print("Loading sales data...")

df = pd.read_excel(
    INPUT_FILE,
    sheet_name="Sales"
)

print(f"Records found: {len(df)}")

# =========================================================
# VALIDATE REQUIRED COLUMNS
# =========================================================

required_columns = [
    "Product",
    "Category",
    "Quantity",
    "Final Price",
    "Unit Cost",
    "Total Sale",
    "Total Cost"
]

for column in required_columns:

    if column not in df.columns:
        raise Exception(f"Missing required column: {column}")

# =========================================================
# DATA CLEANING
# =========================================================

print("Cleaning and validating data...")

# Remove null products
df = df.dropna(subset=["Product"])

# Remove negative values
df = df[
    (df["Quantity"] > 0) &
    (df["Final Price"] > 0) &
    (df["Unit Cost"] > 0)
]

# =========================================================
# CALCULATE PROFITABILITY
# =========================================================

print("Calculating profitability metrics...")

df["Profit"] = (
    df["Total Sale"] - df["Total Cost"]
)

df["Margin %"] = (
    (df["Profit"] / df["Total Sale"]) * 100
).round(2)

# =========================================================
# PROFITABILITY SUMMARY BY PRODUCT
# =========================================================

print("Generating profitability summary...")

profitability_summary = (
    df.groupby(["Product", "Category"])
    .agg({
        "Quantity": "sum",
        "Total Sale": "sum",
        "Total Cost": "sum",
        "Profit": "sum"
    })
    .reset_index()
)

profitability_summary["Margin %"] = (
    (
        profitability_summary["Profit"] /
        profitability_summary["Total Sale"]
    ) * 100
).round(2)

# =========================================================
# PROFITABILITY SCORE
# =========================================================

print("Calculating profitability score...")

max_profit = profitability_summary["Profit"].max()
max_quantity = profitability_summary["Quantity"].max()

profitability_summary["Profitability Score"] = (
    (
        (profitability_summary["Profit"] / max_profit) * 0.5
    ) +
    (
        (profitability_summary["Margin %"] / 100) * 0.3
    ) +
    (
        (profitability_summary["Quantity"] / max_quantity) * 0.2
    )
) * 100

profitability_summary["Profitability Score"] = (
    profitability_summary["Profitability Score"]
    .round(2)
)

# =========================================================
# RANKING
# =========================================================

profitability_summary = profitability_summary.sort_values(
    by="Profit",
    ascending=False
)

profitability_summary["Profitability Rank"] = range(
    1,
    len(profitability_summary) + 1
)

# =========================================================
# TOP PRODUCTS
# =========================================================

top_profitable_products = profitability_summary.head(10)

# =========================================================
# LOW MARGIN PRODUCTS
# =========================================================

low_margin_products = profitability_summary.sort_values(
    by="Margin %",
    ascending=True
).head(10)

# =========================================================
# AUTOMATIC ALERTS
# =========================================================

print("Generating automatic alerts...")

alerts = []

for _, row in profitability_summary.iterrows():

    alert = ""

    if row["Margin %"] < 5:
        alert = "URGENT REVIEW"

    elif row["Margin %"] < 10:
        alert = "LOW MARGIN"

    elif row["Profit"] > profitability_summary["Profit"].mean() * 1.5:
        alert = "HIGH PERFORMER"

    else:
        alert = "STABLE"

    alerts.append(alert)

profitability_summary["Alert"] = alerts

# =========================================================
# PATTERN DETECTION
# =========================================================

print("Detecting business patterns...")

patterns = []

avg_quantity = profitability_summary["Quantity"].mean()
avg_profit = profitability_summary["Profit"].mean()
avg_margin = profitability_summary["Margin %"].mean()

for _, row in profitability_summary.iterrows():

    product = row["Product"]
    quantity = row["Quantity"]
    profit = row["Profit"]
    margin = row["Margin %"]

    # -----------------------------------------------------
    # DECEPTIVE PRODUCT
    # High sales volume but low margin
    # -----------------------------------------------------

    if quantity > avg_quantity and margin < avg_margin:

        patterns.append({
            "Product": product,
            "Pattern Type": "Deceptive Product",
            "Description": "High sales volume but low profitability",
            "Recommended Action": "Review pricing strategy or reduce costs"
        })

    # -----------------------------------------------------
    # PREMIUM PRODUCT
    # Low sales volume but high profit
    # -----------------------------------------------------

    if quantity < avg_quantity and profit > avg_profit:

        patterns.append({
            "Product": product,
            "Pattern Type": "Premium Product",
            "Description": "Lower sales volume but strong profitability",
            "Recommended Action": "Increase product promotion"
        })

    # -----------------------------------------------------
    # CRITICAL MARGIN
    # -----------------------------------------------------

    if margin < 10:

        patterns.append({
            "Product": product,
            "Pattern Type": "Critical Margin",
            "Description": "Profit margin is critically low",
            "Recommended Action": "Negotiate supplier costs immediately"
        })

    # -----------------------------------------------------
    # STAR PRODUCT
    # -----------------------------------------------------

    if profit > avg_profit * 1.5:

        patterns.append({
            "Product": product,
            "Pattern Type": "Star Product",
            "Description": "Highly profitable product",
            "Recommended Action": "Scale sales and marketing efforts"
        })

patterns_df = pd.DataFrame(patterns)

# =========================================================
# EXECUTIVE CONCLUSIONS
# =========================================================

print("Generating executive conclusions...")

top_product = top_profitable_products.iloc[0]
worst_margin = low_margin_products.iloc[0]

conclusions = [

    {
        "Conclusion":
        f"The most profitable product is "
        f"{top_product['Product']} "
        f"with a total profit of "
        f"${top_product['Profit']:.2f}"
    },

    {
        "Conclusion":
        f"The product with the lowest margin is "
        f"{worst_margin['Product']} "
        f"with a margin of "
        f"{worst_margin['Margin %']:.2f}%"
    },

    {
        "Conclusion":
        "Several products generate high sales "
        "volume but low profitability."
    },

    {
        "Conclusion":
        "The company can improve profitability "
        "by prioritizing high-margin products."
    }
]

executive_conclusions = pd.DataFrame(conclusions)

# =========================================================
# GENERAL STATISTICS
# =========================================================

general_statistics = pd.DataFrame([{

    "Total Sales":
        round(df["Total Sale"].sum(), 2),

    "Total Costs":
        round(df["Total Cost"].sum(), 2),

    "Total Profit":
        round(df["Profit"].sum(), 2),

    "Average Margin %":
        round(df["Margin %"].mean(), 2),

    "Products Analyzed":
        df["Product"].nunique(),

    "Total Transactions":
        len(df)
}])

# =========================================================
# CREATE CHARTS
# =========================================================

print("Generating charts...")

# ---------------------------------------------------------
# TOP PROFITABLE PRODUCTS CHART
# ---------------------------------------------------------

top_chart = top_profitable_products.head(10)

plt.figure(figsize=(12, 6))

plt.bar(
    top_chart["Product"],
    top_chart["Profit"]
)

plt.xticks(rotation=45, ha="right")

plt.title("Top Profitable Products")
plt.xlabel("Product")
plt.ylabel("Profit")

plt.tight_layout()

top_chart_file = "top_products_chart.png"

plt.savefig(top_chart_file)

plt.close()

# ---------------------------------------------------------
# PROFITABILITY BY CATEGORY
# ---------------------------------------------------------

category_profit = (
    profitability_summary
    .groupby("Category")["Profit"]
    .sum()
    .sort_values(ascending=False)
)

plt.figure(figsize=(10, 6))

plt.bar(
    category_profit.index,
    category_profit.values
)

plt.xticks(rotation=30)

plt.title("Profitability by Category")
plt.xlabel("Category")
plt.ylabel("Profit")

plt.tight_layout()

category_chart_file = "category_profit_chart.png"

plt.savefig(category_chart_file)

plt.close()

# =========================================================
# EXPORT TO EXCEL
# =========================================================

print("Exporting final report...")

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    profitability_summary.to_excel(
        writer,
        sheet_name="PROFITABILITY_SUMMARY",
        index=False
    )

    top_profitable_products.to_excel(
        writer,
        sheet_name="TOP_PRODUCTS",
        index=False
    )

    low_margin_products.to_excel(
        writer,
        sheet_name="LOW_MARGIN_PRODUCTS",
        index=False
    )

    patterns_df.to_excel(
        writer,
        sheet_name="DETECTED_PATTERNS",
        index=False
    )

    executive_conclusions.to_excel(
        writer,
        sheet_name="EXECUTIVE_CONCLUSIONS",
        index=False
    )

    general_statistics.to_excel(
        writer,
        sheet_name="GENERAL_STATISTICS",
        index=False
    )

# =========================================================
# LOAD WORKBOOK
# =========================================================

wb = load_workbook(OUTPUT_FILE)

# =========================================================
# STYLES
# =========================================================

header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

# =========================================================
# FORMAT SHEETS
# =========================================================

for ws in wb.worksheets:

    # -----------------------------------------------------
    # FORMAT HEADERS
    # -----------------------------------------------------

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font

    # -----------------------------------------------------
    # AUTO WIDTH
    # -----------------------------------------------------

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except:
                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

# =========================================================
# INSERT CHARTS
# =========================================================

charts_sheet = wb.create_sheet("CHARTS")

img1 = ExcelImage(top_chart_file)
img1.width = 900
img1.height = 450

charts_sheet.add_image(img1, "A1")

img2 = ExcelImage(category_chart_file)
img2.width = 800
img2.height = 400

charts_sheet.add_image(img2, "A28")

# =========================================================
# SAVE FILE
# =========================================================

wb.save(OUTPUT_FILE)

# =========================================================
# CONSOLE OUTPUT
# =========================================================

print("\n====================================")
print("PROFITABILITY ANALYSIS COMPLETED")
print("====================================\n")

print("TOP PROFITABLE PRODUCTS:\n")

for _, row in top_profitable_products.iterrows():

    print(
        f"{row['Product']} | "
        f"Profit: ${row['Profit']:.2f} | "
        f"Margin: {row['Margin %']:.2f}%"
    )

print("\nLOW MARGIN PRODUCTS:\n")

for _, row in low_margin_products.iterrows():

    print(
        f"{row['Product']} | "
        f"Margin: {row['Margin %']:.2f}%"
    )

print("\nGenerated file:")
print(OUTPUT_FILE)
